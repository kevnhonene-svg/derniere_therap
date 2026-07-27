from django.contrib.auth import authenticate, login, logout
from django.core.exceptions import ValidationError
from django.db.models import F, Q
from django.http import HttpResponse
from io import BytesIO
from rest_framework import status
from rest_framework.decorators import api_view, authentication_classes, permission_classes
from rest_framework.permissions import AllowAny
from dos_app.commande.models import Commande, QuotaBillet
from dos_app.comptes.models import ConfigurationApplication, Invite, TableGala
from dos_app.comptes.serializers import (
    ConfigurationApplicationSerializer,
    InviteSerializer,
    TableGalaSerializer,
    config_to_dict,
    invite_to_dict,
    table_to_dict,
)
from dos_app.comptes.utils import error, json_body, success
from dos_app.message.models import MessageConversation
from dos_app.stock.models import Boisson


def is_superadmin(request):
    user = request._request.user
    session = request._request.session
    return (user.is_authenticated and user.is_superuser) or session.get('role') == 'superadmin'


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def configuration(request):
    config, _ = ConfigurationApplication.objects.get_or_create(pk=1)
    return success({'configuration': config_to_dict(config)})


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def login_billet(request):
    data = json_body(request)
    code = str(data.get('code_billet', '')).strip()

    if not code:
        return error('Veuillez entrer votre code billet.')

    try:
        invite = Invite.objects.select_related('table').get(code_billet__iexact=code, actif=True)
    except Invite.DoesNotExist:
        return error('Code billet invalide.', status.HTTP_404_NOT_FOUND)

    role = 'protocole' if invite.est_protocole else 'client'

    request._request.session['invite_id'] = invite.id
    request._request.session['role'] = role
    return success({'role': role, 'invite': invite_to_dict(invite)})


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def login_admin(request):
    data = json_body(request)
    username = data.get('username', '')
    password = data.get('password', '')
    user = authenticate(request._request, username=username, password=password)
    if not user or not user.is_superuser:
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)
    login(request._request, user)
    request._request.session['role'] = 'superadmin'
    request._request.session['admin_username'] = user.username
    return success({'role': 'superadmin', 'user': {'username': user.username}})


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def logout_view(request):
    request._request.session.flush()
    logout(request._request)
    return success()


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def me(request):
    role = 'anonymous'
    payload = {}
    user = request._request.user
    session = request._request.session
    if user.is_authenticated and user.is_superuser:
        role = 'superadmin'
        payload['user'] = {'username': user.username}
    elif session.get('role') == 'superadmin':
        role = 'superadmin'
        payload['user'] = {'username': session.get('admin_username', 'Superadmin')}
    elif session.get('invite_id'):
        try:
            invite = Invite.objects.select_related('table').get(pk=session['invite_id'])
            role = session.get('role', 'client')
            payload['invite'] = invite_to_dict(invite)
        except Invite.DoesNotExist:
            session.flush()
    payload['role'] = role
    return success(payload)


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def tables_disponibles(request):
    tables = TableGala.objects.filter(active=True)
    return success({'tables': [table_to_dict(table) for table in tables if not table.est_pleine]})


@api_view(['GET', 'POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def tables_admin(request):
    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        tables = TableGala.objects.all()
        return success({'tables': TableGalaSerializer(tables, many=True).data})

    serializer = TableGalaSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    table = serializer.save()
    return success({'table': table_to_dict(table)}, status.HTTP_201_CREATED)


@api_view(['PATCH', 'DELETE'])
@authentication_classes([])
@permission_classes([AllowAny])
def table_admin_detail(request, table_id):
    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    try:
        table = TableGala.objects.get(pk=table_id)
    except TableGala.DoesNotExist:
        return error('Table introuvable.', status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        if table.invites.exists():
            return error(
                (
                    f"La table {table.nom} contient encore {table.places_occupees} invite(s). "
                    "Retirez ou deplacez les invites avant de supprimer cette table."
                )
            )
        nom = table.nom
        table.delete()
        return success({'message': f'Table {nom} supprimee.'})

    serializer = TableGalaSerializer(table, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    table = serializer.save()
    return success({'table': table_to_dict(table)})


@api_view(['GET', 'POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def invites_admin(request):
    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    if request.method == 'GET':
        invites = Invite.objects.select_related('table').all()
        return success({'invites': InviteSerializer(invites, many=True).data})

    serializer = InviteSerializer(data=request.data)
    serializer.is_valid(raise_exception=True)
    invite = serializer.save()
    try:
        invite.full_clean()
        invite.save()
    except ValidationError as exc:
        invite.delete()
        return error(exc.message_dict if hasattr(exc, 'message_dict') else exc.messages)
    return success({'invite': invite_to_dict(invite)}, status.HTTP_201_CREATED)


@api_view(['PUT', 'PATCH', 'DELETE'])
@authentication_classes([])
@permission_classes([AllowAny])
def invite_admin_detail(request, invite_id):
    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    try:
        invite = Invite.objects.select_related('table').get(pk=invite_id)
    except Invite.DoesNotExist:
        return error('Invite introuvable.', status.HTTP_404_NOT_FOUND)

    if request.method == 'DELETE':
        nom = invite.nom_complet
        code = invite.code_billet
        invite.delete()
        return success({'message': f'Invite {nom} ({code}) supprime.'})

    serializer = InviteSerializer(invite, data=request.data, partial=request.method == 'PATCH')
    serializer.is_valid(raise_exception=True)
    invite = serializer.save()
    try:
        invite.full_clean()
        invite.save()
    except ValidationError as exc:
        return error(exc.message_dict if hasattr(exc, 'message_dict') else exc.messages)
    return success({'invite': invite_to_dict(invite)})


@api_view(['POST'])
@authentication_classes([])
@permission_classes([AllowAny])
def configuration_admin(request):
    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    config, _ = ConfigurationApplication.objects.get_or_create(pk=1)
    serializer = ConfigurationApplicationSerializer(config, data=request.data, partial=True)
    serializer.is_valid(raise_exception=True)
    config = serializer.save()
    return success({'configuration': config_to_dict(config)})


def param_bool(request, name):
    value = request.GET.get(name, '')
    if value == '':
        return None
    return value.lower() in ['1', 'true', 'oui', 'yes']


def date_text(value):
    return value.strftime('%Y-%m-%d %H:%M') if value else ''


def sheet_name(name):
    return name[:31]


FILTER_LABELS = {
    'module': 'Module',
    'categorie_billet': 'Type billet',
    'table_id': 'Table',
    'actif': 'Invite actif',
    'est_protocole': 'Role protocole',
    'has_table': 'Presence table',
    'active': 'Table/boisson active',
    'full': 'Occupation table',
    'stock': 'Stock boisson',
    'statut': 'Statut commande',
    'auteur': 'Auteur message',
    'lu': 'Lecture message',
    'date_from': 'Date debut',
    'date_to': 'Date fin',
}

TICKET_FILLS = {
    'Classique': 'EAF4EF',
    'VIP Couple': 'FCECEC',
    'VIP corps academique': 'FFF4D6',
    'VIP premium': 'EAF0FB',
}


def filter_summary(request):
    items = []
    for key, label in FILTER_LABELS.items():
        value = request.GET.get(key, '').strip()
        if value:
            items.append(f'{label}: {value}')
    return ' | '.join(items) if items else 'Aucun filtre specifique'


def add_rank(rows):
    return [{'N°': index, **row} for index, row in enumerate(rows, start=1)]


def ticket_totals_summary(invites):
    totals = {value: 0 for value, _ in Invite.CATEGORIES_BILLET}
    for invite in invites:
        totals[invite.categorie_billet] = totals.get(invite.categorie_billet, 0) + 1

    parts = [
        f'{label}: {totals.get(value, 0)}'
        for value, label in Invite.CATEGORIES_BILLET
    ]
    parts.append(f'Total: {sum(totals.values())}')
    return 'Totaux billets selon filtre: ' + ' | '.join(parts)


def write_sheet(writer, name, rows, columns, config, filters_text, ticket_totals_text):
    import pandas as pd
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    df = pd.DataFrame(rows, columns=columns)
    safe_name = sheet_name(name)
    startrow = 7
    df.to_excel(writer, sheet_name=safe_name, index=False, startrow=startrow)
    worksheet = writer.sheets[safe_name]
    max_col = max(len(columns), 1)
    last_col = get_column_letter(max_col)

    title_fill = PatternFill('solid', fgColor='183C34')
    subtitle_fill = PatternFill('solid', fgColor='F6F4EE')
    header_fill = PatternFill('solid', fgColor='B89435')
    even_fill = PatternFill('solid', fgColor='FFFFFF')
    odd_fill = PatternFill('solid', fgColor='F8F4EA')
    border_fill = PatternFill('solid', fgColor='EEF5F1')
    thin_border = Border(
        left=Side(style='thin', color='9A8F7C'),
        right=Side(style='thin', color='9A8F7C'),
        top=Side(style='thin', color='9A8F7C'),
        bottom=Side(style='thin', color='9A8F7C'),
    )

    worksheet.merge_cells(f'A1:{last_col}1')
    worksheet.merge_cells(f'A2:{last_col}2')
    worksheet.merge_cells(f'A3:{last_col}3')
    worksheet.merge_cells(f'A4:{last_col}4')
    worksheet.merge_cells(f'A5:{last_col}5')
    worksheet.merge_cells(f'A6:{last_col}6')

    worksheet['A1'] = config.nom_application
    worksheet['A2'] = config.nom_evenement
    worksheet['A3'] = config.sous_titre
    worksheet['A4'] = f'Export: {name}'
    worksheet['A5'] = f'Filtres appliques: {filters_text}'
    worksheet['A6'] = ticket_totals_text

    worksheet['A1'].fill = title_fill
    worksheet['A1'].font = Font(color='FFFFFF', bold=True, size=16)
    worksheet['A2'].fill = title_fill
    worksheet['A2'].font = Font(color='FFFFFF', bold=True, size=13)
    worksheet['A3'].fill = subtitle_fill
    worksheet['A3'].font = Font(color='183C34', italic=True)
    worksheet['A4'].fill = subtitle_fill
    worksheet['A4'].font = Font(color='183C34', bold=True)
    worksheet['A5'].fill = subtitle_fill
    worksheet['A5'].font = Font(color='695B43')
    worksheet['A6'].fill = subtitle_fill
    worksheet['A6'].font = Font(color='183C34', bold=True)

    for row in range(1, 7):
        worksheet.row_dimensions[row].height = 22
        worksheet.cell(row=row, column=1).alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

    header_row = startrow + 1
    for cell in worksheet[header_row]:
        cell.fill = header_fill
        cell.font = Font(color='FFFFFF', bold=True)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border

    ticket_col = None
    for index, column in enumerate(columns, start=1):
        if column == 'Type billet':
            ticket_col = index
            break

    for row_index in range(header_row + 1, worksheet.max_row + 1):
        fill = even_fill if row_index % 2 == 0 else odd_fill
        ticket_fill = None
        if ticket_col:
            ticket_value = worksheet.cell(row=row_index, column=ticket_col).value
            color = TICKET_FILLS.get(ticket_value)
            if color:
                ticket_fill = PatternFill('solid', fgColor=color)
        for column_index in range(1, max_col + 1):
            cell = worksheet.cell(row=row_index, column=column_index)
            cell.fill = ticket_fill or fill
            cell.alignment = Alignment(vertical='top', wrap_text=True)
            cell.border = thin_border

    worksheet.freeze_panes = worksheet['A9']
    worksheet.auto_filter.ref = f'A{header_row}:{last_col}{worksheet.max_row}'
    for column_index in range(1, max_col + 1):
        worksheet.cell(row=7, column=column_index).fill = border_fill
        worksheet.cell(row=7, column=column_index).border = thin_border

    for index, column in enumerate(df.columns, start=1):
        values = [str(column)] + [str(value) for value in df[column].fillna('').tolist()]
        width = min(max(len(value) for value in values) + 2, 42)
        worksheet.column_dimensions[get_column_letter(index)].width = width


def filtered_invites(request):
    qs = Invite.objects.select_related('table').all()
    search = request.GET.get('search', '').strip()
    categorie = request.GET.get('categorie_billet', '')
    table_id = request.GET.get('table_id', '')
    has_table = request.GET.get('has_table', '')
    actif = param_bool(request, 'actif')
    protocole = param_bool(request, 'est_protocole')
    if search:
        qs = qs.filter(
            Q(nom__icontains=search)
            | Q(postnom__icontains=search)
            | Q(prenom__icontains=search)
            | Q(code_billet__icontains=search)
            | Q(telephone__icontains=search)
            | Q(email__icontains=search)
            | Q(table__nom__icontains=search)
        )
    if categorie:
        qs = qs.filter(categorie_billet=categorie)
    if table_id:
        qs = qs.filter(table_id=table_id)
    if has_table == 'oui':
        qs = qs.filter(table__isnull=False)
    if has_table == 'non':
        qs = qs.filter(table__isnull=True)
    if actif is not None:
        qs = qs.filter(actif=actif)
    if protocole is not None:
        qs = qs.filter(est_protocole=protocole)
    return qs


def filtered_tables(request):
    qs = TableGala.objects.all()
    search = request.GET.get('search', '').strip()
    active = param_bool(request, 'active')
    if search:
        qs = qs.filter(nom__icontains=search)
    if active is not None:
        qs = qs.filter(active=active)
    full = request.GET.get('full', '')
    if full:
        tables = [table for table in qs if table.est_pleine == (full == 'oui')]
        return tables
    return qs


def filtered_boissons(request):
    qs = Boisson.objects.all()
    search = request.GET.get('search', '').strip()
    active = param_bool(request, 'active')
    stock = request.GET.get('stock', '')
    if search:
        qs = qs.filter(Q(nom__icontains=search) | Q(categorie__icontains=search) | Q(description__icontains=search))
    if active is not None:
        qs = qs.filter(actif=active)
    if stock == 'disponible':
        qs = qs.filter(actif=True, quantite_stock__gt=0)
    if stock == 'rupture':
        qs = qs.filter(quantite_stock__lte=0)
    if stock == 'faible':
        qs = qs.filter(quantite_stock__lte=F('seuil_alerte'))
    return qs


def filtered_commandes(request):
    qs = Commande.objects.select_related('invite', 'invite__table').prefetch_related('lignes__boisson').all()
    statut = request.GET.get('statut', '')
    categorie = request.GET.get('categorie_billet', '')
    table_id = request.GET.get('table_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '').strip()
    if statut:
        qs = qs.filter(statut=statut)
    if categorie:
        qs = qs.filter(invite__categorie_billet=categorie)
    if table_id:
        qs = qs.filter(invite__table_id=table_id)
    if date_from:
        qs = qs.filter(cree_le__date__gte=date_from)
    if date_to:
        qs = qs.filter(cree_le__date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(invite__nom__icontains=search)
            | Q(invite__postnom__icontains=search)
            | Q(invite__prenom__icontains=search)
            | Q(invite__code_billet__icontains=search)
            | Q(note_client__icontains=search)
            | Q(note_protocole__icontains=search)
        )
    return qs


def filtered_messages(request):
    qs = MessageConversation.objects.select_related('invite', 'invite__table').all()
    auteur = request.GET.get('auteur', '')
    lu = param_bool(request, 'lu')
    invite_id = request.GET.get('invite_id', '')
    date_from = request.GET.get('date_from', '')
    date_to = request.GET.get('date_to', '')
    search = request.GET.get('search', '').strip()
    if auteur:
        qs = qs.filter(auteur=auteur)
    if lu is not None:
        qs = qs.filter(lu=lu)
    if invite_id:
        qs = qs.filter(invite_id=invite_id)
    if date_from:
        qs = qs.filter(cree_le__date__gte=date_from)
    if date_to:
        qs = qs.filter(cree_le__date__lte=date_to)
    if search:
        qs = qs.filter(
            Q(contenu__icontains=search)
            | Q(invite__nom__icontains=search)
            | Q(invite__postnom__icontains=search)
            | Q(invite__prenom__icontains=search)
            | Q(invite__code_billet__icontains=search)
        )
    return qs


@api_view(['GET'])
@authentication_classes([])
@permission_classes([AllowAny])
def export_xlsx(request):
    import pandas as pd

    if not is_superadmin(request):
        return error('Acces reserve au superadmin.', status.HTTP_403_FORBIDDEN)

    module = request.GET.get('module', 'all')
    config, _ = ConfigurationApplication.objects.get_or_create(pk=1)
    filters_text = filter_summary(request)
    output = BytesIO()
    generated = []

    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        invites = list(filtered_invites(request))
        tables = list(filtered_tables(request))
        boissons = list(filtered_boissons(request))
        quotas = list(QuotaBillet.objects.all())
        commandes = list(filtered_commandes(request))
        messages = list(filtered_messages(request))
        ticket_totals_text = ticket_totals_summary(invites)

        if module in ['all', 'invites']:
            rows = add_rank([{
                'Nom complet': invite.nom_complet,
                'Telephone': invite.telephone,
                'Code billet': invite.code_billet,
                'Type billet': invite.get_categorie_billet_display(),
                'Table assignee': invite.table.nom if invite.table else 'Sans table',
                'Protocole': 'Oui' if invite.est_protocole else 'Non',
                'Actif': 'Oui' if invite.actif else 'Non',
                'Cree le': date_text(invite.cree_le),
            } for invite in invites])
            write_sheet(writer, 'Invites', rows, ['N°', 'Nom complet', 'Telephone', 'Code billet', 'Type billet', 'Table assignee', 'Protocole', 'Actif', 'Cree le'], config, filters_text, ticket_totals_text)
            generated.append(('Invites', len(rows)))

        if module in ['all', 'tables']:
            rows = add_rank([{
                'Table': table.nom,
                'Places': table.nombre_places,
                'Occupees': table.places_occupees,
                'Restantes': table.places_restantes,
                'Pleine': 'Oui' if table.est_pleine else 'Non',
                'Active': 'Oui' if table.active else 'Non',
                'Cree le': date_text(table.cree_le),
            } for table in tables])
            write_sheet(writer, 'Tables', rows, ['N°', 'Table', 'Places', 'Occupees', 'Restantes', 'Pleine', 'Active', 'Cree le'], config, filters_text, ticket_totals_text)
            generated.append(('Tables', len(rows)))

        if module in ['all', 'boissons']:
            rows = add_rank([{
                'Nom': boisson.nom,
                'Categorie': boisson.categorie,
                'Description': boisson.description,
                'Prix indicatif': boisson.prix_indicatif,
                'Stock': boisson.quantite_stock,
                'Seuil alerte': boisson.seuil_alerte,
                'Disponible': 'Oui' if boisson.est_disponible else 'Non',
                'Stock faible': 'Oui' if boisson.stock_faible else 'Non',
                'Active': 'Oui' if boisson.actif else 'Non',
                'Cree le': date_text(boisson.cree_le),
            } for boisson in boissons])
            write_sheet(writer, 'Boissons', rows, ['N°', 'Nom', 'Categorie', 'Description', 'Prix indicatif', 'Stock', 'Seuil alerte', 'Disponible', 'Stock faible', 'Active', 'Cree le'], config, filters_text, ticket_totals_text)
            generated.append(('Boissons', len(rows)))

        if module in ['all', 'quotas']:
            rows = add_rank([{
                'Type billet': quota.get_categorie_billet_display(),
                'Nombre bouteilles': quota.nombre_bouteilles,
                'Actif': 'Oui' if quota.actif else 'Non',
            } for quota in quotas])
            write_sheet(writer, 'Quotas', rows, ['N°', 'Type billet', 'Nombre bouteilles', 'Actif'], config, filters_text, ticket_totals_text)
            generated.append(('Quotas', len(rows)))

        if module in ['all', 'commandes']:
            rows = add_rank([{
                'Nom complet': commande.invite.nom_complet,
                'Code billet': commande.invite.code_billet,
                'Type billet': commande.invite.get_categorie_billet_display(),
                'Table': commande.invite.table.nom if commande.invite.table else 'Sans table',
                'Statut': commande.get_statut_display(),
                'Total bouteilles': commande.total_bouteilles,
                'Note client': commande.note_client,
                'Note protocole': commande.note_protocole,
                'Cree le': date_text(commande.cree_le),
                'Mise a jour': date_text(commande.mise_a_jour),
            } for commande in commandes])
            write_sheet(writer, 'Commandes', rows, ['N°', 'Nom complet', 'Code billet', 'Type billet', 'Table', 'Statut', 'Total bouteilles', 'Note client', 'Note protocole', 'Cree le', 'Mise a jour'], config, filters_text, ticket_totals_text)
            ligne_rows = []
            for commande in commandes:
                for ligne in commande.lignes.all():
                    ligne_rows.append({
                        'Commande ID': commande.id,
                        'Invite': commande.invite.nom_complet,
                        'Boisson': ligne.boisson.nom,
                        'Categorie boisson': ligne.boisson.categorie,
                        'Quantite': ligne.quantite,
                        'Statut commande': commande.get_statut_display(),
                    })
            ligne_rows = add_rank(ligne_rows)
            write_sheet(writer, 'Details commandes', ligne_rows, ['N°', 'Commande ID', 'Invite', 'Boisson', 'Categorie boisson', 'Quantite', 'Statut commande'], config, filters_text, ticket_totals_text)
            generated.append(('Commandes', len(rows)))

        if module in ['all', 'messages']:
            rows = add_rank([{
                'Nom complet': message.invite.nom_complet,
                'Code billet': message.invite.code_billet,
                'Table': message.invite.table.nom if message.invite.table else 'Sans table',
                'Auteur': message.get_auteur_display(),
                'Lu': 'Oui' if message.lu else 'Non',
                'Cree le': date_text(message.cree_le),
            } for message in messages])
            write_sheet(writer, 'Messages', rows, ['N°', 'Nom complet', 'Code billet', 'Table', 'Auteur', 'Lu', 'Cree le'], config, filters_text, ticket_totals_text)
            generated.append(('Messages', len(rows)))

        summary_rows = add_rank([{'Feuille': name, 'Lignes exportees': count} for name, count in generated])
        write_sheet(writer, 'Resume', summary_rows, ['N°', 'Feuille', 'Lignes exportees'], config, filters_text, ticket_totals_text)

    output.seek(0)
    filename = f"export_admin_{module}.xlsx"
    response = HttpResponse(output.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response
